'''
Detail: Creating the block storage using VIPR using REST API
Created By: Vijay Kumar Singh
Date:16/05/2016
Last Modified On: 14/06/2016
Parameter:
    Logger_Path: Give the logger location for the script
                 Example: "c:\temp"
    EXCEL_FILE_PATH: Excel File provided by the customer for creation of the BLOCK
                 Example:
                        Its inside the script folder
    VIPR_CONTROLLER: The ip of the VIPR controller or the DNS name if local machine is able to resolve it.
                 Example:
                 10.31.83.129
    VIPR_Port :The port of the VIPR usually 1433. 

    VIPR_USER_NAME:The authorised user for the creation of the Block Storage
                 Example:
                 demo@demo.local
    VIPR_PASSWORD:The password for the VIPR_USER_NAME
'''


import time
import xlrd
import requests
import logging
import requests
from requests import session
import queue
import json

from openpyxl import load_workbook
##Uncomment the below lines if you are getting the certificate warning.
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)






'''Global Variables for a particular VIPR host'''
global viprHost
global viprPort
global login
login='/login'
global logout
global session
logout='/logout'
global token
q=queue.Queue(maxsize=0)    ##Queue will be used for Accepted Request confirmation
    
'''
Module Name :getToken(user,password)
user        :Authorised user name for the Vipr
password    :Password
Description : This will return the token [x-sds-token] that will be used for further request.
'''
def getToken(user,password):
    url = makeURL(login)
    headers = {'ACCEPT': 'application/json'}
    session = requests.Session()
    response = session.get(url, auth=(user, password), verify=False, headers=headers) 
    logging.info(response)
    if response.status_code != 200:
        #If invalid credentials are entered then it will return a text/html format
        if 'text/html' in response.headers['Content-Type']:
            err = "Invalid username or password"
        else:
            error_json = json.loads(response.text)
            err = error_json["details"]
        raise Exception(err)
    if 'x-sds-auth-token' not in response.headers:
        raise Exception("Invalid Login")
    token = response.headers['x-sds-auth-token']
    return token


'''
Module Name  :getHeader(token,contentType,xml)
token        : token is the x-sds-auth-token
contentType  :Default application/json however we can pass it also
xml          :Default xml is false. we can pass it true if we want header in xml
'''
def getHeader(token, contentType='application/json', xml=False):
    if (xml):
        headers = {'Content-Type': contentType, 'ACCEPT': 'application/xml, application/octet-stream'}
    else:
        headers = {'Content-Type': contentType, 'ACCEPT': 'application/json, application/octet-stream'}
    headers['x-sds-auth-token'] = token
    return headers





'''
Module Name   :httpRequest(requestMethod,uri,token,requstBody,contentType,xml
requestMethod : GET,PUT,POST are the valid set of method
token         :token required to build a request
'''

def httpRequest(requestMethod,uri,token,requestBody=None,contentType="application/json",xml=False):
    url=makeURL(uri)
    header=getHeader(token,contentType,xml)
    session = requests.Session()
    if(requestMethod=='GET'):
        response = session.get(url, verify=False, headers=header)
    elif(requestMethod=='POST'):
        response = session.post(url,data=requestBody,verify=False, headers=header)
    elif requestMethod == 'PUT':
        response = session.put(url, data=requestBody, headers=header, verify=False)
    else:
        raise Exception("Invalid request method: " + requestMethod)
    if response.status_code == requests.codes['ok'] or response.status_code == 202 or response.status_code==200:
            logging.debug("Response: %s" % response.text)
            return response
    else:
        logging.error("Request failed: %s" % response.status_code)
        if response.status_code == 401:
            # 401 response is html, so not parsing response
            error_details = "Unauthorized"
        elif 'text/html' in response.headers['Content-Type']:
            root = ET.fromstring(response.text)
            print(response.text)
            error_details = root.find("head/title").text
        else:
            error_json = json.loads(response.text)
            logging.info(error_json)
            if "details" in error_json:
                error_details = error_json["details"]
            else:
                error_details = response.reason
        raise Exception("%s: %s" % (str(response.status_code), error_details))
    
'''
Module Name: releaseToken
token      : token value to be released
importance : This module is important as we have to release token
'''  
def releaseToken(token):
    status=httpRequest("GET",logout,token)
    if(status.status_code==200):
        logging.info("Successfully logout")
    
    
'''
Module Name: makeURL(uri)
uri        : uri will be like "/login","/logout"
'''

def makeURL(uri):
    return "https://{0}:{1}{2}".format(viprHost, viprPort, uri)


'''Detail will be containing all the value required for the operation
   detail["Array"] will  tell the array column position
   detail["VirtualPool"] will tell the Virtual Pool Column position
   '''
def mapping_field(row_labels):
    detail={}
    if("Array" in row_labels):
        detail["Array"]=row_labels.index("Array")
    else:
        raise Exception("Array column is not found")
    if("Virtual Pool" in row_labels):
        detail["VirtualPool"]=row_labels.index("Virtual Pool")
    else:
        raise Exception("Virtual Pool Column is not found")
    if("Project name" in row_labels):
        detail["ProjectName"]=row_labels.index("Project name")
    else:
        raise Exception("Project name column not found")
    if("LUN Name" in row_labels):
        detail["Name"]=row_labels.index("LUN Name")
    else:
        raise Exception("LUN Name column not found")
    if("RP CG" in row_labels):
        detail["ConsistencyGroup"]=row_labels.index("RP CG")
    else:
        raise Exception("RP CG column not found")
    if("Capacity\nGB" in row_labels):
        detail["Capacity"]=row_labels.index("Capacity\nGB")
    else:
        raise Exception("Capacity\nGB column not found")
    if("Status" in row_labels):
        detail["Status"]=row_labels.index("Status")
    else:
        raise Exception("Status column not found")
    return detail



'''
Module Name: preparePayLoad(count,name,project,size,varray,vpool,consistencygrp)
Last Modify:11th-June-2016
'''
def preparePayLoad(count,volumeName,size,projectUrn,varrayUrn,vpoolUrn,consistencygrp=None):
    data = {}
    data['name'] = volumeName
    data['size']=size
    data['count']=count
    data['project']=projectUrn
    data['varray']=varrayUrn
    data['vpool']=vpoolUrn
    data['consistency_group']=consistencygrp
    json_data = json.dumps(data)
    return json_data

'''
Module Name:findUrnId(name,category,token)
Created By : Vijay Kumar Singh
Last Modify: 11th-June-2016
name       : holding the name the of the entity for which we want to search URN
category   : category valid set virtualArray,virtualPool,projectName & consistencyGroup
token      : will be used for submitting the request

'''


def findUrnId(name,category,token):
    if(category=="virtualArray"):
        uri="/vdc/varrays/search?name="+name
        response=httpRequest("GET",uri,token)
        res=json.loads(response.text)
        return (res["resource"][0]["id"])
    if(category=="virtualPool"):
        uri="/block/vpools/search?name="+name
        response=httpRequest("GET",uri,token)
        res=json.loads(response.text)
        return (res["resource"][0]["id"])
    if(category=="projectName"):
        uri="/projects/search?name="+name
        response=httpRequest("GET",uri,token)
        res=json.loads(response.text)
        return (res["resource"][0]["id"])
    if(category=="consistencyGroup"):
        uri="/block/consistency-groups/search?name="+name
        response=httpRequest("GET",uri,token)
        res=json.loads(response.text)
        return (res["resource"][0]["id"])



'''
Module Name    :waitTillRequestSuccessful(responseData)
Description    :Checking the status of progress of submitted request 
'''
def waitTillRequestSuccessfulOrFail(data):
    taskId=data["taskId"]
    volumeId=data["volumeId"]
    taskState=data["state"]
    token=data["token"]
    ##/block/volumes/{id}/tasks/{op_id}/
    apiUrl="/block/volumes/"+volumeId+"/tasks/"+taskId+"/"
    responseData=httpRequest("GET",apiUrl,token,requestBody=None,contentType="application/json",xml=False)
    respondData=json.loads(responseData.text)
    respondData["token"]=token
    respondData["taskId"]=taskId
    respondData["volumeId"]=volumeId
    print(".",end="\r")
    if(taskState != "error" and taskState !="ready"):
        return waitTillRequestSuccessfulOrFail(respondData)
    else:
        print("\r")
        return taskState
        
    
    
    

    

'''
Module Name: submittingExcelDetails(excelFilePath,token)
Created By     : Vijay Kumar Singh
Last Modify    :11th-June-2016
excelFilePath  : Absolute path for excel file containing the excel sheet
token          :token for submitting the request

Description: This function will take the excel file parse it and submit the detail
             one by one to Vipr for volume creation
'''

    
def submittingExcelDetails(excelFilePath,token):
    virtualArrayMap={}
    virtualPoolMap={}
    consistencyGrpMap={}
    projectMap={}
    currentsheet=0
    count='1'   ;
    outputexcelpath=excelFilePath[:excelFilePath.index(".xls")]+"-OutputFile"+excelFilePath[excelFilePath.index(".xls"):]
    wb = load_workbook(excelFilePath)
    readExcelWorkbook=xlrd.open_workbook(excelFilePath)
    sheets=readExcelWorkbook.sheets()
    for sheet in sheets:
        if(currentsheet!=sheet):
            currentsheet=sheet
            ws = wb.get_sheet_by_name(currentsheet.name)
            row_labels=currentsheet.row_values(0,0,currentsheet.ncols)
            indexdetail=mapping_field(row_labels)
        for eachrow in range(1,currentsheet.nrows):
            statuscell = ws.cell(row=eachrow+1,column=indexdetail["Status"]+1) ## While writing range is not starting from 0.
            #statuscell.value="Successful"
            
            virtualArrayName=currentsheet.cell(eachrow,indexdetail["Array"]).value
            virtualPoolName=currentsheet.cell(eachrow,indexdetail["VirtualPool"]).value
            consistencyGroupName=currentsheet.cell(eachrow,indexdetail["ConsistencyGroup"]).value
            projectGroupName=currentsheet.cell(eachrow,indexdetail["ProjectName"]).value
            status=currentsheet.cell(eachrow,indexdetail["Status"]).value
            size=str(currentsheet.cell(eachrow,indexdetail["Capacity"]).value)+"GB"
            volumeName=currentsheet.cell(eachrow,indexdetail["Name"]).value
            print("Creating LUN: "+volumeName,end='\r')
            if(virtualArrayName not in virtualArrayMap):
                try:
                    virtualArrayMap[virtualArrayName]=findUrnId(virtualArrayName,"virtualArray",token)
                except Exception as e:
                    statuscell.value="Virtual Array Column"+str(e)
                    print("Failed:"+str(e)+"\r")
                    continue
            if(virtualPoolName not in virtualPoolMap):
                try:
                    virtualPoolMap[virtualPoolName]=findUrnId(virtualPoolName,"virtualPool",token)
                except Exception as e:
                    statuscell.value="Virtual Pool Column"+str(e)
                    print("Failed:"+str(e)+"\r")
                    continue    
            if(consistencyGroupName not in consistencyGrpMap and consistencyGroupName!=42):
                try:
                    consistencyGrpMap[consistencyGroupName]=findUrnId(consistencyGroupName,"consistencyGroup",token)
                except Exception as e:
                    statuscell.value="Consistency Group Column"+str(e)
                    print("Failed:"+str(e)+"\r")
                    continue
            if(projectGroupName not in projectMap):
                try:
                    projectMap[projectGroupName]=findUrnId(projectGroupName,"projectName",token)
                except Exception as e:
                    statuscell.value="Project Group Column"+str(e)
                    print("Failed:"+str(e)+"\r")
                    continue
            if(consistencyGroupName!=42):
                payload=preparePayLoad(count,volumeName,size,projectMap[projectGroupName],virtualArrayMap[virtualArrayName],virtualPoolMap[virtualPoolName],consistencyGrpMap[consistencyGroupName])
            else:
                payload=preparePayLoad(count,volumeName,size,projectMap[projectGroupName],virtualArrayMap[virtualArrayName],virtualPoolMap[virtualPoolName])
            try:
                response=httpRequest("POST","/block/volumes",token,requestBody=payload,contentType="application/json",xml=False)
                responseData=json.loads(response.text)
                if(responseData["task"][0]["state"]=="ready"):
                    statuscell.value="Successful"
                    continue
                elif(responseData["task"][0]["state"]=="error"):
                    statuscell.value=responseData["task"][0]["service_error"]["description"]
                    print("Failed:"+statuscell.value+"\r")
                    continue
                else:
                    print("Successfully Submitted for Creation \r");
                    dataForQueue=responseData
                    dataForQueue["statuscell"]=statuscell
                    dataForQueue["token"]=token
                    dataForQueue["taskId"]=responseData["task"][0]["id"]
                    dataForQueue["state"]=responseData["task"][0]["state"]
                    dataForQueue["volumeId"]=responseData["task"][0]["resource"]["id"]
                    q.put(dataForQueue)
                    continue
            except Exception as e:
                statuscell.value="Failed to create Volume due to:"+str(e)
                print("Failed:"+str(e)+"\r")
                continue
        print("All records are processed and submitted.Please wait for sometime.We are generating report at")
        print(outputexcelpath)
        while not q.empty():
            responseData=q.get()
            status=waitTillRequestSuccessfulOrFail(responseData)
            if(status == "ready"):
                responseData["statuscell"].value="Successful"
            else:
                responseData["statuscell"].value="Error"
    wb.save(outputexcelpath)
    print("Successfully Processed Please check the output of records");

            


                          


                          




def main(userName,password,excelFilePath):
    ##Establishing the connection between VIPR and host
    token=getToken(userName,password)
    ##Parsing Excel file and creating block volumes
    submittingExcelDetails(excelFilePath,token)
    ##Releasing the token after fetching the whole record.
    response=httpRequest("GET","/user/whoami",token)
    releaseToken(token)

if (__name__ == "__main__"):
    viprHost=input("Enter the VIPR host")
    viprPort=input("Enter the PORT for vipr Host")
    Logger_Path=input("Enter the loggerpath")
    userName=input("Enter the authorised User")
    password=input("Enter the password")
    excelFilePath=input("Enter the excel File Path")
    LOG_FILE_NAME=time.strftime("%d_%m_%Y")
    logging.basicConfig(filename=Logger_Path+"\\%s"%LOG_FILE_NAME,level=logging.DEBUG,format='%(asctime)s %(levelname)s %(message)s')
    logging.info("Started logging new run")
    main(userName,password,excelFilePath)

    






